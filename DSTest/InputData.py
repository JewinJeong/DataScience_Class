
import pymysql
import pandas as pd
from openpyxl import load_workbook


df_from_excel = pd.read_excel('db_score.xlsx', sheet_name='Sheet1')
data = [tuple(x) for x in df_from_excel.values]

conn = pymysql.connect(host = 'localhost', user = 'root', password= 'jj8575412', db = 'db_score')
curs = conn.cursor()

sql = """insert into score(sno, attendance, homework, discussion, midterm, final, score, grade) 
        values (%s,%s,%s,%s,%s,%s,%s,%s)"""
data = [tuple(x) for x in df_from_excel.values]

curs.executemany(sql,data)

conn.commit()

conn.close()



# try:
#     with conn.cursor() as curs:
#         sql = """insert into score(sno, attendance, homework, discussion, midterm, final, score, grade)
#             values(%s,%s,%s,%s,%s,%s,%s,%s)"""
#         wb = load_workbook('./db_score.xlsx', data_only=True)
#         ws = wb['Sheet1']
#
#         iter_rows = iter(ws.rows)
#         next(iter_rows)
#         for row in iter_rows:
#             curs.execute(sql,(row[0].data,row[1].data,row[2].data,row[3].data,row[4].data,row[5].data,row[6].value,data[7].value))
#         conn.commit()
# finally:
#     conn.close()
#     wb.close()
